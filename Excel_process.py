from openpyxl import load_workbook

# 加载 Excel 工作簿
wb = load_workbook(filename='小雅导入题目模板v2.1.2200.xlsx')

# 选择一个工作表
ws = wb['Sheet1']

# 遍历工作表中的所有行和列
for row in ws.iter_rows():
    for cell in row:
        print(cell.value)

# 或者按行读取数据
for row in ws.values:
    print(row)
